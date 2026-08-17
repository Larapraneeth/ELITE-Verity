from pathlib import Path
from openpyxl import load_workbook


def extract_excel_content(excel_path):

    excel_path = Path(excel_path)

    workbook = load_workbook(
        excel_path,
        data_only=True
    )

    pages = []

    for worksheet in workbook.worksheets:

        rows = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            continue

        headers = []

        for value in rows[0]:

            if value is None:
                headers.append("")
            else:
                headers.append(
                    str(value).strip()
                )

        sheet_lines = []

        for row_number, row in enumerate(
            rows[1:],
            start=2
        ):

            values = []

            for column_index, value in enumerate(
                row
            ):

                if value is None:
                    continue

                if column_index < len(headers):
                    header = headers[
                        column_index
                    ]
                else:
                    header = (
                        f"Column {column_index + 1}"
                    )

                if not header:
                    header = (
                        f"Column {column_index + 1}"
                    )

                values.append(
                    f"{header}: {value}"
                )

            if values:

                sheet_lines.append(
                    f"Row {row_number}: "
                    + " | ".join(values)
                )

        if not sheet_lines:
            continue

        text = (
            f"Sheet: {worksheet.title}\n\n"
            + "\n".join(sheet_lines)
        )

        pages.append(
            {
                "page": worksheet.title,
                "text": text,
                "sheet": worksheet.title
            }
        )

    workbook.close()

    return pages